from service.file_reader.base import ReaderBase
from PIL import Image
import io
from openpyxl import load_workbook
import base64


class XlsxReader(ReaderBase):
    """
    日志渠道
    """

    channel: str = "xlsx"

    def __init__(self):
        super().__init__()

    def read_bytes(self):
        content = []
        wb = load_workbook(filename=self.file_stream)

        for sheet_name in wb.sheetnames:
            worksheet = wb[sheet_name]
            # 第一步：遍历所有行，找出最大有效列数（基于非空单元格）
            max_col_idx = 0
            rows_data = []

            for row in worksheet.iter_rows(
                values_only=False
            ):  # 注意：先用 cell 对象判断是否为空
                # 找到该行最后一个非空单元格的索引
                last_col = 0
                row_values = []
                for cell in row:
                    row_values.append(cell.value)
                    if cell.value is not None:
                        last_col = cell.column  # column 是整数（A=1, B=2...）
                if last_col > max_col_idx:
                    max_col_idx = last_col
                # 只保留至少有一个非空值的行
                if any(v is not None for v in row_values):
                    rows_data.append(row_values)

            # 如果没有数据，跳过
            if not rows_data or max_col_idx == 0:
                continue

            # 第二步：截断每行到最大有效列
            trimmed_data = []
            for row_vals in rows_data:
                trimmed_row = row_vals[:max_col_idx]  # 截取前 max_col_idx 列
                trimmed_data.append(trimmed_row)

            if trimmed_data:
                content.append({"type": "table", "ext": ".table", "data": trimmed_data})

            # 第三步：读取图片（保持不变）
            for image in worksheet._images:
                image_data = image._data()
                img = Image.open(io.BytesIO(image_data))
                image_ext = img.format.lower()
                img_data = base64.b64encode(image_data).decode("utf-8")
                content.append(
                    {
                        "type": "image",
                        "ext": image_ext,
                        "data": f"data:image/{image_ext};base64,{img_data}",
                    }
                )

        return content
